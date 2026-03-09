import streamlit as st
import os
import json
import shutil
import logging
from pathlib import Path
from datetime import datetime
from typing import Optional, Dict, List, Tuple
from copy import copy
import pandas as pd
import openpyxl
import requests
import tempfile

# ========================================
# Box API クラス (元のロジックを維持)
# ========================================
BOX_API_BASE = 'https://api.box.com/2.0'
BOX_UPLOAD_URL = 'https://upload.box.com/api/2.0/files/content'

class BoxUploader:
    def __init__(self, access_token: str):
        self.access_token = access_token
        self.headers = {
            'Authorization': f'Bearer {access_token}'
        }
        self.folder_cache: Dict[str, str] = {}
    
    def get_folder_items(self, folder_id: str) -> List[Dict]:
        try:
            url = f'{BOX_API_BASE}/folders/{folder_id}/items'
            params = {'limit': 1000, 'fields': 'id,name,type'}
            response = requests.get(url, headers=self.headers, params=params)
            response.raise_for_status()
            return response.json().get('entries', [])
        except requests.exceptions.RequestException as e:
            st.error(f"フォルダ情報の取得に失敗しました: {e}")
            return []
    
    def find_subfolder(self, parent_folder_id: str, folder_name: str) -> Optional[str]:
        cache_key = f"{parent_folder_id}:{folder_name}"
        if cache_key in self.folder_cache:
            return self.folder_cache[cache_key]
        
        items = self.get_folder_items(parent_folder_id)
        for item in items:
            if item['type'] == 'folder' and item['name'] == folder_name:
                self.folder_cache[cache_key] = item['id']
                return item['id']
        return None
    
    def create_subfolder(self, parent_folder_id: str, folder_name: str) -> Optional[str]:
        existing_id = self.find_subfolder(parent_folder_id, folder_name)
        if existing_id:
            return existing_id
        
        try:
            url = f'{BOX_API_BASE}/folders'
            data = {
                'name': folder_name,
                'parent': {'id': parent_folder_id}
            }
            response = requests.post(url, headers=self.headers, json=data)
            
            if response.status_code == 409:
                return self.find_subfolder(parent_folder_id, folder_name)
            
            response.raise_for_status()
            folder_id = response.json()['id']
            cache_key = f"{parent_folder_id}:{folder_name}"
            self.folder_cache[cache_key] = folder_id
            return folder_id
            
        except requests.exceptions.RequestException as e:
            st.error(f"フォルダ作成エラー '{folder_name}': {e}")
            return None

    def get_file_id_in_folder(self, folder_id: str, file_name: str) -> Optional[str]:
        items = self.get_folder_items(folder_id)
        for item in items:
            if item['type'] == 'file' and item['name'] == file_name:
                return item['id']
        return None

    def upload_file_version(self, file_id: str, file_path: Path, file_name: str) -> bool:
        try:
            upload_url = f"https://upload.box.com/api/2.0/files/{file_id}/content"
            with open(file_path, 'rb') as f:
                files = {'file': (file_name, f)}
                response = requests.post(upload_url, headers=self.headers, files=files)
            response.raise_for_status()
            return True
        except requests.exceptions.RequestException as e:
            st.error(f"ファイル更新エラー '{file_name}': {e}")
            return False

    def upload_file(self, folder_id: str, file_path: Path, file_name: str) -> bool:
        try:
            attributes = {
                'name': file_name,
                'parent': {'id': folder_id}
            }
            with open(file_path, 'rb') as f:
                files = {
                    'attributes': (None, json.dumps(attributes)),
                    'file': (file_name, f)
                }
                response = requests.post(BOX_UPLOAD_URL, headers=self.headers, files=files)
            
            if response.status_code == 409:
                existing_file_id = self.get_file_id_in_folder(folder_id, file_name)
                if existing_file_id:
                    return self.upload_file_version(existing_file_id, file_path, file_name)
                else:
                    return False
            
            response.raise_for_status()
            return True
            
        except requests.exceptions.RequestException as e:
            st.error(f"アップロードエラー '{file_name}': {e}")
            return False

# ========================================
# Excel処理関数
# ========================================
def copy_sheet_style(source_sheet, target_sheet):
    for col_char, dimension in source_sheet.column_dimensions.items():
        target_sheet.column_dimensions[col_char].width = dimension.width
        target_sheet.column_dimensions[col_char].hidden = dimension.hidden
    for row_idx, dimension in source_sheet.row_dimensions.items():
        target_sheet.row_dimensions[row_idx].height = dimension.height
    for row in source_sheet.iter_rows():
        for cell in row:
            new_cell = target_sheet.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                new_cell.font = copy(cell.font)
                new_cell.border = copy(cell.border)
                new_cell.fill = copy(cell.fill)
                new_cell.number_format = copy(cell.number_format)
                new_cell.protection = copy(cell.protection)
                new_cell.alignment = copy(cell.alignment)
    for merged_range in source_sheet.merged_cells.ranges:
        target_sheet.merge_cells(str(merged_range))

def process_and_upload(input_path, def_path, token, parent_id, suffix, temp_dir):
    
    # 定義ファイルの読み込み準備
    wb_def_source = None
    if def_path and os.path.exists(def_path):
        try:
            wb_def_source = openpyxl.load_workbook(def_path)
        except Exception as e:
            st.warning(f"定義ファイルの読み込みに失敗: {e}")

    # データ読み込み
    try:
        df_original = pd.read_excel(input_path, header=1)
    except Exception as e:
        st.error(f"エクセル読み込みエラー: {e}")
        return

    # フィルタリング
    df_filtered = df_original.copy()
    exclude_mask = pd.Series(False, index=df_filtered.index)
    if '出現回数' in df_filtered.columns:
        exclude_mask |= (df_filtered['出現回数'] == 1)
    if '案件名' in df_filtered.columns:
        exclude_mask |= (df_filtered['案件名'].astype(str).str.contains('SALE情報', na=False))
    
    df_filtered = df_filtered[~exclude_mask]
    
    project_id_col = '案件ID'
    shop_url_col = '店舗URL'

    if project_id_col not in df_filtered.columns or shop_url_col not in df_filtered.columns:
        st.error(f"必須カラム（{project_id_col}, {shop_url_col}）が見つかりません。")
        return

    grouped = df_filtered.groupby([shop_url_col, project_id_col])
    total_groups = len(grouped)
    
    st.info(f"処理対象: {total_groups} 件のファイルを作成・アップロードします。")
    
    uploader = BoxUploader(token)
    progress_bar = st.progress(0)
    status_text = st.empty()
    
    success_count = 0
    failure_count = 0
    
    # 処理ループ
    for i, ((shop_url, project_id), group_df) in enumerate(grouped):
        safe_shop_url = str(shop_url).replace('/', '').replace('\\', '')
        safe_project_id = str(project_id).replace('/', '').replace('\\', '')
        file_name = f"[{safe_project_id}]{suffix}.xlsx"
        
        # 一時ファイル作成
        temp_shop_dir = os.path.join(temp_dir, safe_shop_url)
        os.makedirs(temp_shop_dir, exist_ok=True)
        target_path = os.path.join(temp_shop_dir, file_name)
        
        # ファイル生成
        shutil.copy(input_path, target_path)
        wb = openpyxl.load_workbook(target_path)
        ws = wb.active
        
        valid_rows = set(group_df.index + 3)
        last_data_row = 2 + len(df_original)
        
        for r in range(last_data_row, 2, -1):
            if r not in valid_rows:
                ws.delete_rows(r)
        
        # 定義シート追加
        if wb_def_source and "データ定義" in wb_def_source.sheetnames:
            try:
                target_sheet = wb.create_sheet("データ定義")
                copy_sheet_style(wb_def_source["データ定義"], target_sheet)
            except Exception:
                pass
        
        wb.save(target_path)
        wb.close()
        
        # Boxアップロード
        status_text.text(f"Uploading: {safe_shop_url} / {file_name}")
        shop_folder_id = uploader.create_subfolder(parent_id, safe_shop_url)
        
        if shop_folder_id:
            if uploader.upload_file(shop_folder_id, Path(target_path), file_name):
                success_count += 1
            else:
                failure_count += 1
        else:
            failure_count += 1
            
        progress_bar.progress((i + 1) / total_groups)

    if wb_def_source:
        wb_def_source.close()

    st.success("処理完了！")
    st.write(f"✅ 成功: {success_count} ファイル")
    if failure_count > 0:
        st.error(f"❌ 失敗: {failure_count} ファイル")

# ========================================
# Streamlit UI メイン
# ========================================
def main():
    st.set_page_config(page_title="Box Uploader", layout="wide")
    st.title("📦 Box Uploader Web")
    st.markdown("エクセルを集計・分割してBoxへ自動アップロードします。")

    # サイドバー設定
    with st.sidebar:
        st.header("設定")
        box_token = st.text_input("Box Developer Token", type="password", help="Box開発者コンソールで発行したトークン")
        parent_folder_id = st.text_input("Box 親フォルダID", value="364342582851")
        file_suffix = st.text_input("ファイル名の末尾", value="9月モニター")

    # ファイルアップロード
    col1, col2 = st.columns(2)
    with col1:
        uploaded_main = st.file_uploader("集計元エクセル (xlsx)", type=["xlsx"])
    with col2:
        uploaded_def = st.file_uploader("定義エクセル (xlsx, 任意)", type=["xlsx"])

    # 実行ボタン
    if st.button("処理開始", type="primary"):
        if not box_token:
            st.error("Boxトークンを入力してください。")
            return
        if not uploaded_main:
            st.error("集計元エクセルをアップロードしてください。")
            return

        # 一時ディレクトリで処理
        with tempfile.TemporaryDirectory() as temp_dir:
            # アップロードされたファイルを一時保存
            main_path = os.path.join(temp_dir, "source.xlsx")
            with open(main_path, "wb") as f:
                f.write(uploaded_main.getbuffer())
            
            def_path = None
            if uploaded_def:
                def_path = os.path.join(temp_dir, "def.xlsx")
                with open(def_path, "wb") as f:
                    f.write(uploaded_def.getbuffer())

            # 処理実行
            process_and_upload(main_path, def_path, box_token, parent_folder_id, file_suffix, temp_dir)

if __name__ == "__main__":
    main()
